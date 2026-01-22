from __future__ import annotations

import base64
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP, getcontext
from io import BytesIO
from typing import Any

from flask import Flask, render_template, request, send_file
import matplotlib
matplotlib.use("Agg")  # Использовать без GUI
from matplotlib import pyplot as plt
from matplotlib.figure import Figure
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font


app = Flask(__name__)

getcontext().prec = 28


@dataclass(frozen=True)
class MortgageResult:
    monthly_payment_rub: Decimal
    overpayment_rub: Decimal
    overpayment_percent: Decimal
    total_paid_rub: Decimal


def _to_decimal(value: str) -> Decimal:
    value = (value or "").strip().replace(" ", "").replace(",", ".")
    if value == "":
        raise ValueError("Пустое значение")
    try:
        return Decimal(value)
    except InvalidOperation as exc:
        raise ValueError("Некорректное число") from exc


def calculate_mortgage(
    home_price: Decimal,
    down_payment: Decimal,
    years: Decimal,
    annual_rate_percent: Decimal,
) -> tuple[MortgageResult, list[dict[str, Decimal]]]:
    if home_price <= 0:
        raise ValueError("Стоимость жилья должна быть больше 0")
    if down_payment < 0:
        raise ValueError("Первоначальный взнос не может быть отрицательным")
    if down_payment >= home_price:
        raise ValueError("Первоначальный взнос должен быть меньше стоимости жилья")
    if years <= 0:
        raise ValueError("Срок кредита должен быть больше 0")
    if annual_rate_percent < 0:
        raise ValueError("Ставка не может быть отрицательной")

    principal = home_price - down_payment

    months = (years * Decimal(12))
    if months != months.to_integral_value():
        raise ValueError("Срок кредита должен быть целым числом лет (например, 15)")
    months_i = int(months)

    if annual_rate_percent == 0:
        monthly = (principal / Decimal(months_i)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
        total_paid = monthly * Decimal(months_i)
    else:
        r = (annual_rate_percent / Decimal(100)) / Decimal(12)  # месячная ставка
        one_plus_r_pow_n = (Decimal(1) + r) ** Decimal(months_i)
        monthly_raw = principal * (r * one_plus_r_pow_n) / (one_plus_r_pow_n - Decimal(1))
        monthly = monthly_raw.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
        total_paid = (monthly * Decimal(months_i)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)

    overpayment_rub = (total_paid - principal).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    overpayment_percent = (overpayment_rub / principal * Decimal(100)).quantize(
        Decimal("0.01"), rounding=ROUND_HALF_UP
    )

    # Построение графика платежей
    schedule: list[dict[str, Decimal]] = []
    balance = principal

    if annual_rate_percent == 0:
        for month_index in range(1, months_i + 1):
            if month_index == months_i:
                principal_part = balance
                payment = principal_part
            else:
                principal_part = monthly
                payment = monthly
            interest_part = Decimal("0.00")
            balance = (balance - principal_part).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
            if balance < Decimal("0"):
                balance = Decimal("0.00")
            schedule.append(
                {
                    "month": Decimal(month_index),
                    "payment": payment,
                    "interest": interest_part,
                    "principal": principal_part,
                    "balance": balance,
                }
            )
    else:
        r = (annual_rate_percent / Decimal(100)) / Decimal(12)
        for month_index in range(1, months_i + 1):
            interest_part = (balance * r).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
            principal_part = (monthly - interest_part).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
            if month_index == months_i:
                principal_part = balance
                payment = (interest_part + principal_part).quantize(
                    Decimal("0.01"), rounding=ROUND_HALF_UP
                )
            else:
                payment = monthly
            balance = (balance - principal_part).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
            if balance < Decimal("0"):
                balance = Decimal("0.00")
            schedule.append(
                {
                    "month": Decimal(month_index),
                    "payment": payment,
                    "interest": interest_part,
                    "principal": principal_part,
                    "balance": balance,
                }
            )

    return (
        MortgageResult(
            monthly_payment_rub=monthly,
            overpayment_rub=overpayment_rub,
            overpayment_percent=overpayment_percent,
            total_paid_rub=total_paid,
        ),
        schedule,
    )


def _fmt_rub(amount: Decimal) -> str:
    # 1 234 567.89
    s = f"{amount:.2f}"
    int_part, frac = s.split(".")
    chunks = []
    while int_part:
        chunks.append(int_part[-3:])
        int_part = int_part[:-3]
    return f"{' '.join(reversed(chunks))}.{frac}"


def _build_schedule_chart(schedule: list[dict[str, Decimal]]) -> str:
    """Создает график платежей и возвращает base64 строку изображения."""
    months = [int(row["month"]) for row in schedule]
    payments = [float(row["payment"]) for row in schedule]
    interests = [float(row["interest"]) for row in schedule]
    principals = [float(row["principal"]) for row in schedule]
    balances = [float(row["balance"]) for row in schedule]

    # Настройка темной темы для графика
    plt.style.use("dark_background")
    fig = Figure(figsize=(12, 8), facecolor="#0b1220")
    ax1 = fig.add_subplot(2, 1, 1)
    ax2 = fig.add_subplot(2, 1, 2)

    # Первый график: платежи, проценты и тело кредита
    ax1.plot(months, payments, label="Платёж", color="#6ea8ff", linewidth=2)
    ax1.fill_between(months, interests, 0, alpha=0.4, color="#ff6b6b", label="Проценты")
    ax1.fill_between(
        months,
        [i + p for i, p in zip(interests, principals)],
        interests,
        alpha=0.4,
        color="#6ef0c2",
        label="Тело кредита",
    )
    ax1.set_xlabel("Месяц", color="#93a4c7", fontsize=11)
    ax1.set_ylabel("Сумма, ₽", color="#93a4c7", fontsize=11)
    ax1.set_title("Структура платежей", color="#e9eefc", fontsize=13, fontweight="bold", pad=15)
    ax1.legend(loc="upper right", facecolor="#0f1a2e", edgecolor=(1.0, 1.0, 1.0, 0.1), labelcolor="#e9eefc")
    ax1.grid(True, alpha=0.2, color="#93a4c7")
    ax1.set_facecolor("#0f1a2e")
    border_color = (1.0, 1.0, 1.0, 0.1)
    ax1.spines["bottom"].set_color(border_color)
    ax1.spines["top"].set_color(border_color)
    ax1.spines["right"].set_color(border_color)
    ax1.spines["left"].set_color(border_color)
    ax1.tick_params(colors="#93a4c7")

    # Второй график: остаток долга
    ax2.plot(months, balances, label="Остаток долга", color="#6ea8ff", linewidth=2)
    ax2.fill_between(months, balances, 0, alpha=0.3, color="#6ea8ff")
    ax2.set_xlabel("Месяц", color="#93a4c7", fontsize=11)
    ax2.set_ylabel("Остаток, ₽", color="#93a4c7", fontsize=11)
    ax2.set_title("Динамика остатка долга", color="#e9eefc", fontsize=13, fontweight="bold", pad=15)
    ax2.legend(loc="upper right", facecolor="#0f1a2e", edgecolor=(1.0, 1.0, 1.0, 0.1), labelcolor="#e9eefc")
    ax2.grid(True, alpha=0.2, color="#93a4c7")
    ax2.set_facecolor("#0f1a2e")
    border_color = (1.0, 1.0, 1.0, 0.1)
    ax2.spines["bottom"].set_color(border_color)
    ax2.spines["top"].set_color(border_color)
    ax2.spines["right"].set_color(border_color)
    ax2.spines["left"].set_color(border_color)
    ax2.tick_params(colors="#93a4c7")

    fig.tight_layout(pad=2.0)

    # Сохранение в base64
    buf = BytesIO()
    fig.savefig(buf, format="png", facecolor="#0b1220", edgecolor="none", dpi=100, bbox_inches="tight")
    buf.seek(0)
    img_base64 = base64.b64encode(buf.read()).decode("utf-8")
    buf.close()
    plt.close(fig)

    return img_base64


def _build_schedule_xlsx(
    *,
    title: str,
    home_price: Decimal,
    down_payment: Decimal,
    years: Decimal,
    annual_rate_percent: Decimal,
    result: MortgageResult,
    schedule: list[dict[str, Decimal]],
) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "График"

    header_font = Font(bold=True)

    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:E1")

    meta = [
        ("Стоимость жилья, ₽", home_price),
        ("Первоначальный взнос, ₽", down_payment),
        ("Срок, лет", years),
        ("Ставка, % годовых", annual_rate_percent),
        ("Ежемесячный платеж, ₽", result.monthly_payment_rub),
        ("Полная сумма, ₽", result.total_paid_rub),
        ("Переплата, ₽", result.overpayment_rub),
        ("Переплата, %", result.overpayment_percent),
    ]
    row = 3
    for label, value in meta:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = header_font
        ws[f"B{row}"] = float(value)
        row += 1

    start_row = row + 1
    headers = ["Месяц", "Платёж, ₽", "Проценты, ₽", "Тело, ₽", "Остаток, ₽"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=col, value=h)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for i, entry in enumerate(schedule, start=1):
        r = start_row + i
        ws.cell(row=r, column=1, value=int(entry["month"]))
        ws.cell(row=r, column=2, value=float(entry["payment"]))
        ws.cell(row=r, column=3, value=float(entry["interest"]))
        ws.cell(row=r, column=4, value=float(entry["principal"]))
        ws.cell(row=r, column=5, value=float(entry["balance"]))

    # Форматы
    for r in range(start_row + 1, start_row + 1 + len(schedule)):
        ws.cell(row=r, column=1).alignment = Alignment(horizontal="left")
        for c in (2, 3, 4, 5):
            ws.cell(row=r, column=c).number_format = '#,##0.00'
            ws.cell(row=r, column=c).alignment = Alignment(horizontal="right")

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 16

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out


@app.get("/")
def index() -> str:
    return render_template(
        "index.html",
        form={
            "home_price": "",
            "down_payment": "",
            "years": "",
            "annual_rate_percent": "",
        },
        is_installment=False,
        result=None,
        schedule=None,
        error=None,
    )


@app.post("/calculate")
def calculate() -> str:
    mode = request.form.get("mode", "credit")
    is_installment = mode == "installment"

    form = {
        "home_price": request.form.get("home_price", ""),
        "down_payment": request.form.get("down_payment", ""),
        "years": request.form.get("years", ""),
        "annual_rate_percent": request.form.get("annual_rate_percent", ""),
    }

    try:
        annual_rate = Decimal(0) if is_installment else _to_decimal(form["annual_rate_percent"])

        result, schedule = calculate_mortgage(
            home_price=_to_decimal(form["home_price"]),
            down_payment=_to_decimal(form["down_payment"]),
            years=_to_decimal(form["years"]),
            annual_rate_percent=annual_rate,
        )
        view_result: dict[str, Any] = {
            "monthly_payment_rub": _fmt_rub(result.monthly_payment_rub),
            "overpayment_rub": _fmt_rub(result.overpayment_rub),
            "overpayment_percent": f"{result.overpayment_percent:.2f}",
            "total_paid_rub": _fmt_rub(result.total_paid_rub),
        }

        schedule_view = [
            {
                "month": int(row["month"]),
                "payment": _fmt_rub(row["payment"]),
                "interest": _fmt_rub(row["interest"]),
                "principal": _fmt_rub(row["principal"]),
                "balance": _fmt_rub(row["balance"]),
            }
            for row in schedule
        ]

        chart_base64 = _build_schedule_chart(schedule)

        return render_template(
            "index.html",
            form=form,
            is_installment=is_installment,
            result=view_result,
            schedule=schedule_view,
            chart_base64=chart_base64,
            error=None,
        )
    except ValueError as exc:
        return render_template(
            "index.html",
            form=form,
            is_installment=is_installment,
            result=None,
            schedule=None,
            chart_base64=None,
            error=str(exc),
        )


@app.post("/export_excel")
def export_excel():
    mode = request.form.get("mode", "credit")
    is_installment = mode == "installment"

    form = {
        "home_price": request.form.get("home_price", ""),
        "down_payment": request.form.get("down_payment", ""),
        "years": request.form.get("years", ""),
        "annual_rate_percent": request.form.get("annual_rate_percent", ""),
    }

    home_price = _to_decimal(form["home_price"])
    down_payment = _to_decimal(form["down_payment"])
    years = _to_decimal(form["years"])
    annual_rate = Decimal(0) if is_installment else _to_decimal(form["annual_rate_percent"])

    result, schedule = calculate_mortgage(
        home_price=home_price,
        down_payment=down_payment,
        years=years,
        annual_rate_percent=annual_rate,
    )

    title = "Рассрочка" if is_installment else "Ипотека"
    xlsx = _build_schedule_xlsx(
        title=f"{title}: график платежей",
        home_price=home_price,
        down_payment=down_payment,
        years=years,
        annual_rate_percent=annual_rate,
        result=result,
        schedule=schedule,
    )

    return send_file(
        xlsx,
        as_attachment=True,
        download_name="mortgage_schedule.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        max_age=0,
    )


if __name__ == "__main__":
    app.run(host="127.0.0.1", port=5000, debug=True)

